# -*- coding: utf-8 -*-
"""
Created on Tue Sep  3 13:58:40 2019

@author: JakeC
"""
# Imports
import sqlite3
import openpyxl
        
# Creating and connecting to the database        
connection = sqlite3.connect("database.db")
cursor = connection.cursor()
    
# Opening the excel sheets
wb1 = openpyxl.load_workbook('inspections.xlsx')   
wb2 = openpyxl.load_workbook('violations.xlsx')

sheet1 = wb1['inspections']
sheet2 = wb2['violations']

inspections_sql = "DROP TABLE IF EXISTS inspections;"
cursor.execute(inspections_sql)

violations_sql = "DROP TABLE IF EXISTS violations;"
cursor.execute(violations_sql)


# Creating inspections table
inspections_sql = """
CREATE TABLE inspections (
    activity_date date,
    employee_id varchar(12),
    facility_address varchar(120),
    facility_city varchar(60),
    facility_id varchar(12),
    facility_name varchar(100),
    facility_state varchar(2),
    facility_zip varchar(10),
    grade text,
    owner_id varchar(12),
    owner_name varchar(100),
    pe_description text,
    program_element_pe integer(4),
    program_name varchar(100),
    program_status text,
    record_id varchar(12),
    score integer(3),
    serial_number varchar(15),
    service_code integer(5),
    service_description text
);"""
        
cursor.execute(inspections_sql)

# Creating violations table
violations_sql = """
CREATE TABLE violations (
    points integer(2),
    serial_number varchar(15),
    violation_code varchar(5),
    violation_description text,
    violation_status text
);"""

cursor.execute(violations_sql)

 # Inserting the values into the database       
insert_inspections = """
INSERT INTO inspections (
        activity_date,
        employee_id,
        facility_address,
        facility_city,
        facility_id,
        facility_name,
        facility_state,
        facility_zip,
        grade,
        owner_id,
        owner_name,
        pe_description,
        program_element_pe,
        program_name,
        program_status,
        record_id,
        score,
        serial_number,
        service_code,
        service_description
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"""

insert_violations = """
INSERT INTO violations (
        points,
        serial_number,
        violation_code,
        violation_description,
        violation_status
) VALUES (?, ?, ?, ?, ?);"""

#Looping through all the rows in the excel sheet starting from row 2       
for row in sheet1.iter_rows(min_row = 2):
    cursor.execute(insert_inspections, [row[i].value for i in range(20)])
    
for row in sheet2.iter_rows(min_row = 2):
    cursor.execute(insert_violations, [row[i].value for i in range(5)])

# Saves and closes the connection
connection.commit()
connection.close()