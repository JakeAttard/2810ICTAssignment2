# -*- coding: utf-8 -*-
"""
Created on Sat Sep 21 22:38:21 2019

@author: JakeC
"""
# Imports
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Font
import sqlite3

# Connecting to the database
connection = sqlite3.connect("database.db")
cursor = connection.cursor()

# Creating the excel workbook
wb = Workbook()
wbs1 = wb.active

# Adding the excel titles and bolding the headings as shown in the task sheet.
wbs1.title = "Violation Types"
wbs1["A1"] ="Code"
wbs1["A1"].font = Font(bold = True)
wbs1["B1"] = "Description"
wbs1["B1"].font = Font(bold = True)
wbs1["C1"] = "Count"
wbs1["C1"].font = Font(bold = True)

print(wb.sheetnames)

# SQL command to get the data from the sqlite database
sqlSelectQuery = """
SELECT violation_code,
    violation_description,
    COUNT(violation_code) FROM violations GROUP BY violation_code ORDER BY violation_code"""

cursor.execute(sqlSelectQuery)

# Fetching all the data from the database in a variable called violationTypes.
violationTypes = cursor.fetchall()

# Looping through all the data and putting it in the excel.
for i in violationTypes:
    wbs1.append(i)

# Shows the total sum of count
counter = 0
for row in wbs1.iter_rows(min_row = 2):
    count = row[2].value
    counter += count

lastRow = wbs1.max_row

# Adding a heading for the total sum on line 118 in row B
wbs1["B118"] = "Total Sum"

wbs1['C' + str(lastRow + 1) ] = int(counter)

# Saving all the data in a excel sheet called ViolationTypes.xlsx
wb.save("ViolationTypes.xlsx")